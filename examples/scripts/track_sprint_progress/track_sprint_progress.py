"""
Python script to analyze Jira sprint progress using the pyJiraCli tool.
"""
################################################################################
# Imports
################################################################################
import sys
import os
import json
import subprocess

from enum import IntEnum
from datetime import datetime
from openpyxl import Workbook, load_workbook

################################################################################
# Configuration
################################################################################

__author__ = ""
__date__ = datetime.now().strftime('%d.%m.%Y')

###############################################
# Project configurations
###############################################

# The board for which the sprints shall be analyzed
BOARD = "BSP_BOARD: COMPONENT_1"

# the server profile which shall be used to connect to jira by the pyJiraCLi tool
SERVER_PROFILE = 'pyJiraCli_profile'

# on which excel table the data will be stored
EXCEL_TABLE = "Tabelle1"

# The files where the data will be stored
EXCEL_FILE = f"C:/project/{BOARD.replace(' ','_').replace(':', '')}_JiraProjectProgress.xlsx"
MISSING_DATA_FILE = f"C:/project/{BOARD.replace(' ','_').replace(':', '')}_MissingTicketData.txt"

################################################################################
# Variables
################################################################################

TMP_FILE = ".\\tmp.json"

HEADER_ROW = 5
CONCLUSION_ROW = 6
INITIAL_ROW = 8
DATE_CELL = "L3"

TICKETS_NO_TIME_TRACKING = "no_time_tracking"
TICKETS_NO_SPENT_TIME = "no_spent_time"


class Formulas:
    """Contains static methods to return Excel formulas as strings."""

    @staticmethod
    def timebuffer_f(row: int) -> str:
        """
        Return the formula to calculate time buffer for a given row.

        Args:
            row (int): The row number.

        Returns:
            str: The formula for the time buffer.
        """
        return f"= P{row} - (O{row} + N{row} + M{row})"

    @staticmethod
    def processed_time_units_f(row: int) -> str:
        """
        Return the formula to calculate processed time units for a given row.

        Args:
            row (int): The row number.

        Returns:
            str: The formula for the processed time units.
        """
        return f"=IF(O{row} > 0, O{row} + Q{row}, 0)"

    @staticmethod
    def efficiency_f(row: int) -> str:
        """
        Return the formula to calculate efficiency for a given row.

        Args:
            row (int): The row number.

        Returns:
            str: The formula for efficiency.
        """
        return f"=IF(O{row} > 0, P{row}/O{row}, 0)"

    @staticmethod
    def progress_f(row: int) -> str:
        """
        Return the formula to calculate progress for a given row.

        Args:
            row (int): The row number.

        Returns:
            str: The formula for progress.
        """
        return f"=(R{row}+N{row}/2)/(R{row}+M{row}+N{row})"

    @staticmethod
    def sum_f(length: int, col: str) -> str:
        """
        Return the formula to calculate the sum of a column over a given length.

        Args:
            length (int): The number of rows.
            col (str): The column letter.

        Returns:
            str: The formula for the sum.
        """
        return f"=SUM({col}{INITIAL_ROW}:{col}{INITIAL_ROW + (length - 1)})"


class Cells(IntEnum):
    """Enum for the different cell positions in the Excel sheet."""
    SPRINT_NAME = 0
    START_DATE = 1
    END_DATE = 2
    DELTA_DAYS = 3
    TIME_EST_IDLE = 4
    TIME_REM_OPEN = 5
    TIME_SPENT_DONE = 6
    ORIGINAL_EST = 7
    TIME_BUFFER = 8
    PROCESSED_TIME_UNITS = 9
    EFFICIENCY = 10
    PROGRESS = 11


class Conclusions(IntEnum):
    """Enum for the conclusion rows in the Excel sheet."""
    TOTAL_TIME_EST_IDLE = 13
    TOTAL_TIME_REM_OPEN = 14
    TOTAL_TIME_SPENT_DONE = 15
    TOTAL_ORIGINAL_EST = 16
    TOTAL_TIME_BUFFER = 17
    TOTAL_PROCESSED_TIME_UNITS = 18


DICT_KEYS = {
    Cells.SPRINT_NAME: 'sprint_name',
    Cells.START_DATE: 'startDate',
    Cells.END_DATE: 'endDate',
    Cells.DELTA_DAYS: 'delta_days',
    Cells.TIME_EST_IDLE: 'time_est_idle',
    Cells.TIME_REM_OPEN: 'time_remaining_open',
    Cells.TIME_SPENT_DONE: 'time_spent_done',
    Cells.ORIGINAL_EST: 'total_time_est'
}

EXCEL_COLS = {
    Cells.SPRINT_NAME: 'L',
    Cells.START_DATE: 'U',
    Cells.END_DATE: 'V',
    Cells.DELTA_DAYS: 'W',
    Cells.TIME_EST_IDLE: 'M',
    Cells.TIME_REM_OPEN: 'N',
    Cells.TIME_SPENT_DONE: 'O',
    Cells.ORIGINAL_EST: 'P',
    Cells.TIME_BUFFER: 'Q',
    Cells.PROCESSED_TIME_UNITS: 'R',
    Cells.EFFICIENCY: 'S',
    Cells.PROGRESS: 'T'
}

CONCLUSION_COLS = {
    Conclusions.TOTAL_TIME_EST_IDLE: EXCEL_COLS.get(Cells.TIME_EST_IDLE),
    Conclusions.TOTAL_TIME_REM_OPEN: EXCEL_COLS.get(Cells.TIME_REM_OPEN),
    Conclusions.TOTAL_TIME_SPENT_DONE: EXCEL_COLS.get(Cells.TIME_SPENT_DONE),
    Conclusions.TOTAL_ORIGINAL_EST: EXCEL_COLS.get(Cells.ORIGINAL_EST),
    Conclusions.TOTAL_TIME_BUFFER: EXCEL_COLS.get(Cells.TIME_BUFFER),
    Conclusions.TOTAL_PROCESSED_TIME_UNITS: EXCEL_COLS.get(Cells.PROCESSED_TIME_UNITS),
}

EXCEL_FORMULAS = {
    Cells.TIME_BUFFER: Formulas.timebuffer_f,
    Cells.PROCESSED_TIME_UNITS: Formulas.processed_time_units_f,
    Cells.EFFICIENCY: Formulas.efficiency_f,
    Cells.PROGRESS: Formulas.progress_f,
}

################################################################################
# Functions
################################################################################


def write_conclusion(sheet, num_sprints):
    """
    Write conclusion data to the Excel file.

    Args:
        sheet (Worksheet): The Excel worksheet object.
        num_sprints (int): The number of sprints.
    """
    for conclusion, col in CONCLUSION_COLS.items():  # pylint: disable=W0612
        cell_ref = f"{col}{CONCLUSION_ROW}"
        data = Formulas.sum_f(num_sprints, col)
        sheet[cell_ref] = data


def write_missing_data_file(tickets_no_timetracking: list, tickets_no_spent_time: list):
    """
    Write missing data information to a file.

    Args:
        tickets_no_timetracking (list): List of tickets with no time estimation.
        tickets_no_spent_time (list): List of closed tickets with no spent time.
    """
    with open(MISSING_DATA_FILE, mode='w', encoding='utf-8') as file:
        file.write("Tickets with no time estimation:\n")
        for (key, url) in tickets_no_timetracking:
            file.write(f"{key} ({url})\n")

        file.write("\nClosed tickets with no spent time:\n")
        for (key, url) in tickets_no_spent_time:
            file.write(f"{key} ({url})\n")


def write_excel_file(sheet, board_dict: dict) -> tuple:
    """
    Write the board data to an Excel sheet.

    Args:
        sheet (Worksheet): The Excel worksheet object.
        board_dict (dict): A dictionary containing the board data.

    Returns:
        tuple: Lists of tickets with no time tracking and no spent time.
    """
    tickets_no_time_tracking = []
    tickets_no_spent_time = []

    row_num = INITIAL_ROW

    for sprint in board_dict['sprints']:

        sprint_data = board_dict['sprints'][sprint]

        # Write the data to the corresponding cells
        for key, value in sprint_data.items():
            if key in DICT_KEYS.values():
                col_letter = list(EXCEL_COLS.values())[
                    list(DICT_KEYS.values()).index(key)]
                cell = f"{col_letter}{row_num}"
                sheet[cell] = value

        # Apply formulas to the row
        for cell, formula in EXCEL_FORMULAS.items():
            col_letter = EXCEL_COLS[cell]
            cell_ref = f"{col_letter}{row_num}"
            sheet[cell_ref] = formula(row_num)

        tickets_no_time_tracking.extend(sprint_data[TICKETS_NO_TIME_TRACKING])
        tickets_no_spent_time.extend(sprint_data[TICKETS_NO_SPENT_TIME])

        row_num += 1

    return tickets_no_time_tracking, tickets_no_spent_time


def write_data(board_dict: dict) -> None:
    """
    Write sprint data to an Excel file.

    Args:
        board_dict (dict): A dictionary containing sprint data.
    """
    if not os.path.exists(EXCEL_FILE):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = EXCEL_TABLE
    else:
        workbook = load_workbook(EXCEL_FILE)
        sheet = workbook[EXCEL_TABLE]

    tickets_no_time_tracking, tickets_no_spent_time = write_excel_file(
        sheet, board_dict)
    num_sprints = len(board_dict['sprints'])

    write_conclusion(sheet, num_sprints)
    sheet[DATE_CELL] = __date__

    workbook.save(EXCEL_FILE)

    write_missing_data_file(tickets_no_time_tracking, tickets_no_spent_time)
    print(
        f"Data from JSON file has been successfully written to '{EXCEL_FILE}'")


def get_time_est(issue_data: dict) -> int:
    """
    Get the total estimated time for an issue.

    Args:
        issue_data (dict): A dictionary containing issue data.

    Returns:
        int: The total estimated time in hours.
    """
    time_est = 0

    if 'timeoriginalestimate' in issue_data['fields']:
        if issue_data['fields']['timeoriginalestimate'] is not None:
            time_est = issue_data['fields']['timeoriginalestimate']
    elif 'aggregatetimeoriginalestimate' in issue_data['fields']:
        if issue_data['fields']['aggregatetimeoriginalestimate'] is not None:
            time_est = issue_data['fields']['aggregatetimeoriginalestimate']
    elif 'timeestimate' in issue_data['fields']:
        if issue_data['fields']['timeestimate'] is not None:
            time_est = issue_data['fields']['timeestimate']
    elif 'originalEstimateSeconds' in issue_data['fields']['timetracking']:
        if issue_data['fields']['timetracking']['originalEstimateSeconds'] is not None:
            time_est = issue_data['fields']['timetracking']['originalEstimateSeconds']
    else:
        print(f"No original time estimate found for issue {issue_data['key']}")

    return time_est // 3600


def get_time_spent(issue_data: dict) -> int:
    """
    Get the total time spent for an issue.

    Args:
        issue_data (dict): A dictionary containing issue data.

    Returns:
        int: The total time spent in hours.
    """
    time_spent = 0

    if 'timespent' in issue_data['fields']:
        if issue_data['fields']['timespent'] is not None:
            time_spent = issue_data['fields']['timespent']
    elif 'aggregatetimespent' in issue_data['fields']:
        if issue_data['fields']['aggregatetimespent'] is not None:
            time_spent = issue_data['fields']['aggregatetimespent']
    elif 'timeSpentSeconds' in issue_data['fields']['timetracking']:
        if issue_data['fields']['timetracking']['timeSpentSeconds'] is not None:
            time_spent = issue_data['fields']['timetracking']['timeSpentSeconds']
    else:
        print(f"No original time estimate found for issue {issue_data['key']}")

    return time_spent // 3600


def convert_date_format(date_str: str) -> str:
    """
    Convert a date string from the original format to a new format.

    Args:
        date_str (str): The original date string.

    Returns:
        str: The formatted date string.
    """
    date_format = "%Y-%m-%dT%H:%M:%S.%f%z"
    date_obj = datetime.strptime(date_str, date_format)
    new_format = "%d.%m.%Y"
    formatted_date = date_obj.strftime(new_format)

    return formatted_date


def calculate_days_between_dates(date1_str: str, date2_str: str) -> int:
    """
    Calculate the number of days between two dates.

    Args:
        date1_str (str): The first date string.
        date2_str (str): The second date string.

    Returns:
        int: The number of days between the two dates.
    """
    date_format = "%Y-%m-%dT%H:%M:%S.%f%z"
    date1 = datetime.strptime(date1_str, date_format)
    date2 = datetime.strptime(date2_str, date_format)
    difference = abs(date2 - date1)
    days_between = difference.days

    return days_between


def process_sprint(raw_dict: dict) -> dict:
    """
    Process the raw sprint data into a structured dictionary.

    Args:
        raw_dict (dict): The raw sprint data.

    Returns:
        dict: The processed sprint data.
    """
    processed_dict = {}

    tickets_without_time_tracking = []
    closed_tickets_without_spent_time = []

    raw_sprint_dict = {
        'total_time_est': 0,
        'total_time_spent': 0,
        'total_time_remaining': 0,
        'time_est_open': 0,
        'time_spent_open': 0,
        'time_remaining_open': 0,
        'time_est_idle': 0,
        'time_spent_idle': 0,
        'time_remaining_idle': 0,
        'time_est_closed': 0,
        'time_spent_closed': 0,
        'time_remaining_closed': 0
    }

    for issue in raw_dict['issues']:
        state = issue['fields']['status']['name']
        url_parts = issue['self'].split('/')
        url_parts.remove('')
        url = f"{url_parts[0]}//{url_parts[1]}/browse/{issue['key']}"

        time_est = get_time_est(issue)
        time_spent = get_time_spent(issue)
        time_remaining = time_est - time_spent

        if time_est == 0:
            tickets_without_time_tracking.append((issue['key'], url))

        if state == 'Offen':
            raw_sprint_dict['time_est_idle'] += time_est
            raw_sprint_dict['time_spent_idle'] += time_spent
            raw_sprint_dict['time_remaining_idle'] += time_remaining
        elif state == 'In Arbeit':
            raw_sprint_dict['time_est_open'] += time_est
            raw_sprint_dict['time_spent_open'] += time_spent
            raw_sprint_dict['time_remaining_open'] += time_remaining
        elif state == 'Gelöst':
            raw_sprint_dict['time_est_closed'] += time_est
            raw_sprint_dict['time_spent_closed'] += time_spent
            raw_sprint_dict['time_remaining_closed'] += time_remaining

            if time_spent == 0:
                closed_tickets_without_spent_time.append((issue['key'], url))

        raw_sprint_dict['total_time_est'] += time_est
        raw_sprint_dict['total_time_spent'] += time_spent
        raw_sprint_dict['total_time_remaining'] += time_remaining

    processed_dict = {
        DICT_KEYS.get(Cells.SPRINT_NAME): raw_dict['name'],
        DICT_KEYS.get(Cells.START_DATE): convert_date_format(raw_dict['startdate']),
        DICT_KEYS.get(Cells.END_DATE): convert_date_format(raw_dict['endDate']),
        DICT_KEYS.get(Cells.DELTA_DAYS): calculate_days_between_dates(
            raw_dict['startdate'],
            raw_dict['endDate']),
        DICT_KEYS.get(Cells.TIME_EST_IDLE): raw_sprint_dict['time_est_idle'],
        DICT_KEYS.get(Cells.TIME_REM_OPEN): raw_sprint_dict['time_remaining_open'],
        DICT_KEYS.get(Cells.TIME_SPENT_DONE): raw_sprint_dict['time_spent_closed'] +
        raw_sprint_dict['time_spent_open'],
        DICT_KEYS.get(Cells.ORIGINAL_EST): raw_sprint_dict['total_time_est'],
        TICKETS_NO_TIME_TRACKING: tickets_without_time_tracking,
        TICKETS_NO_SPENT_TIME: closed_tickets_without_spent_time
    }

    return processed_dict


def process_data(board_dict: dict) -> dict:
    """
    Process the raw board data into a structured dictionary.

    Args:
        board_dict (dict): The raw board data.

    Returns:
        dict: The processed board data.
    """
    processed_dict = {
        'board': board_dict['board']
    }

    processed_dict['sprints'] = {}

    for sprint in board_dict['sprints']:
        sprint_dict = process_sprint(board_dict['sprints'][sprint])
        processed_dict['sprints'][sprint] = sprint_dict

    return processed_dict


def get_sprint_data(board: str, profile: str) -> dict:
    """
    Get sprint data for a board using pyJiraCli.

    Args:
        board (str): The board name.
        profile (str): The server profile name.

    Returns:
        dict: The sprint data.
    """
    sprint_dict = {}
    command = f'pyJiraCli --verbose --profile {profile} get_sprints "{board}" --file {TMP_FILE}'

    result = subprocess.run(command, shell=True, check=False)

    if result.returncode == 0:
        try:
            with open(TMP_FILE, mode='r', encoding='utf-8') as file:
                sprint_dict = json.load(file)
            os.remove(TMP_FILE)
        except Exception as e:  # pylint: disable=W0718
            print(e)
            return None

    print(f"Subprocess returned code {result.returncode}")
    return sprint_dict


def get_issue_data(sprint: str, profile: str) -> dict:
    """
    Get issue data for a sprint using pyJiraCli.

    Args:
        sprint (str): The sprint name.
        profile (str): The server profile name.

    Returns:
        dict: The issue data.
    """
    issue_dict = {}
    command = f'pyJiraCli --verbose \
                          --profile {profile} \
                            search "sprint = \'{sprint}\'" \
                          --file {TMP_FILE}'

    result = subprocess.run(command, shell=True, check=False)

    if result.returncode == 0:
        try:
            with open(TMP_FILE, mode='r', encoding='utf-8') as file:
                issue_dict = json.load(file)
            os.remove(TMP_FILE)
        except Exception as e:  # pylint: disable=W0718
            print(e)
            return None

    print(f"Subprocess returned code {result.returncode}")
    return issue_dict


def get_board_data(board: str, profile: str) -> dict:
    """
    Get board data including sprint and issue data using pyJiraCli.

    Args:
        board (str): The board name.
        profile (str): The server profile name.

    Returns:
        dict: The board data.
    """
    board_dict = {}
    sprint_dict = get_sprint_data(board, profile)

    if 'sprints' in sprint_dict:
        board_dict['board'] = board
        board_dict['sprints'] = {}

        for sprint in sprint_dict['sprints']:
            board_dict['sprints'][sprint['name']] = {
                'name': sprint['name'],
                'startdate': sprint['startDate'],
                'endDate': sprint['endDate'],
                'status': sprint['state']
            }

            issue_dict = get_issue_data(sprint['name'], profile)

            if 'issues' in issue_dict:
                board_dict['sprints'][sprint['name']
                                      ]['issues'] = issue_dict['issues']

    return board_dict

################################################################################
# main
################################################################################


def main():
    """
    Main function to execute the script.
    """
    raw_data = get_board_data(BOARD, SERVER_PROFILE)

    if 'board' in raw_data:
        processed_data = process_data(raw_data)
        write_data(processed_data)
    else:
        print("Some error occurred")


################################################################################
# system entry
################################################################################
if __name__ == "__main__":
    sys.exit(main())
